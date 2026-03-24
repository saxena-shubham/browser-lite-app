"""
Browser Lite - Web Content Extraction App
A lightweight tool to extract, process, and export web content from URLs.
Supports YouTube, Instagram, product pages, and general web pages.
AI-powered extraction via Groq and Google Gemini (free tiers available).
"""

import csv
import io
import json
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, parse_qs, urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup
from flask import Flask, render_template, request, jsonify, send_file
from readability import Document

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

# ============================================================
# URL Classification
# ============================================================

def classify_url(url):
    """Classify a URL by platform type."""
    try:
        parsed = urlparse(url)
        domain = parsed.netloc.lower()
        if any(d in domain for d in ['youtube.com', 'youtu.be']):
            return 'youtube'
        elif 'instagram.com' in domain:
            return 'instagram'
        else:
            return 'general'
    except:
        return 'general'


# ============================================================
# YouTube Extraction
# ============================================================

def extract_youtube(url):
    """Extract metadata and transcript from YouTube videos."""
    result = {
        'url': url,
        'type': 'youtube',
        'title': '',
        'description': '',
        'channel': '',
        'duration': '',
        'views': '',
        'upload_date': '',
        'thumbnail': '',
        'transcript': '',
        'tags': '',
        'status': 'success'
    }

    # Primary: use yt-dlp for comprehensive metadata
    try:
        import yt_dlp
        ydl_opts = {
            'quiet': True,
            'no_warnings': True,
            'skip_download': True,
            'no_check_certificates': True,
        }
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=False)
            result['title'] = info.get('title', '')
            result['description'] = (info.get('description', '') or '')[:2000]
            result['channel'] = info.get('channel', info.get('uploader', ''))

            dur = info.get('duration', 0)
            if dur:
                mins, secs = divmod(int(dur), 60)
                hrs, mins = divmod(mins, 60)
                result['duration'] = f"{hrs}:{mins:02d}:{secs:02d}" if hrs else f"{mins}:{secs:02d}"

            views = info.get('view_count', 0)
            result['views'] = f"{views:,}" if views else ''

            ud = info.get('upload_date', '')
            if ud and len(ud) == 8:
                result['upload_date'] = f"{ud[:4]}-{ud[4:6]}-{ud[6:8]}"

            result['thumbnail'] = info.get('thumbnail', '')
            result['tags'] = ', '.join(info.get('tags', [])[:10])
    except Exception as e:
        # Fallback: oEmbed API
        try:
            oembed_url = f"https://www.youtube.com/oembed?url={url}&format=json"
            resp = requests.get(oembed_url, headers=HEADERS, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                result['title'] = data.get('title', '')
                result['channel'] = data.get('author_name', '')
                result['thumbnail'] = data.get('thumbnail_url', '')
            else:
                result['status'] = 'partial'
                result['error'] = f"yt-dlp failed: {str(e)}"
        except Exception as e2:
            result['status'] = 'error'
            result['error'] = f"All methods failed: {str(e)}, {str(e2)}"

    # Transcript extraction
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
        parsed = urlparse(url)
        if 'youtu.be' in parsed.netloc:
            video_id = parsed.path.strip('/')
        else:
            video_id = parse_qs(parsed.query).get('v', [''])[0]

        if video_id:
            transcript_list = YouTubeTranscriptApi.get_transcript(video_id)
            result['transcript'] = ' '.join([t['text'] for t in transcript_list])
    except Exception as e:
        result['transcript_note'] = f"Transcript unavailable: {str(e)[:100]}"

    return result


# ============================================================
# Instagram Extraction
# ============================================================

def extract_instagram(url):
    """Extract content from Instagram posts/reels."""
    result = {
        'url': url,
        'type': 'instagram',
        'title': '',
        'author': '',
        'caption': '',
        'media_url': '',
        'thumbnail': '',
        'post_type': '',
        'status': 'success'
    }

    # Method 1: oEmbed API (works for public posts, no auth needed)
    try:
        oembed_url = f"https://api.instagram.com/oembed?url={url}&maxwidth=640"
        resp = requests.get(oembed_url, headers=HEADERS, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            result['title'] = data.get('title', '')
            result['author'] = data.get('author_name', '')
            result['thumbnail'] = data.get('thumbnail_url', '')
            html = data.get('html', '')
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                result['caption'] = soup.get_text(strip=True)[:1000]
            return result
    except:
        pass

    # Method 2: Scrape Open Graph tags
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(resp.text, 'html.parser')

        og_title = soup.find('meta', attrs={'property': 'og:title'})
        result['title'] = og_title.get('content', '') if og_title else ''

        og_desc = soup.find('meta', attrs={'property': 'og:description'})
        result['caption'] = og_desc.get('content', '')[:1000] if og_desc else ''

        og_image = soup.find('meta', attrs={'property': 'og:image'})
        result['thumbnail'] = og_image.get('content', '') if og_image else ''

        og_type = soup.find('meta', attrs={'property': 'og:type'})
        result['post_type'] = og_type.get('content', '') if og_type else ''
    except Exception as e:
        result['status'] = 'error'
        result['error'] = str(e)

    return result


# ============================================================
# General Web Page Extraction
# ============================================================

def extract_general(url):
    """Extract structured content from any web page."""
    result = {
        'url': url,
        'type': 'general',
        'title': '',
        'meta_description': '',
        'headings': [],
        'main_text': '',
        'links': [],
        'images': [],
        'tables': [],
        'og_data': {},
        'status': 'success'
    }

    try:
        resp = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')

        # Title
        if soup.title and soup.title.string:
            result['title'] = soup.title.string.strip()
        else:
            og_title = soup.find('meta', attrs={'property': 'og:title'})
            result['title'] = og_title.get('content', '') if og_title else ''

        # Meta description
        meta = soup.find('meta', attrs={'name': 'description'})
        if not meta:
            meta = soup.find('meta', attrs={'property': 'og:description'})
        result['meta_description'] = meta.get('content', '') if meta else ''

        # Open Graph data
        for og in soup.find_all('meta', attrs={'property': re.compile(r'^og:')}):
            key = og.get('property', '').replace('og:', '')
            result['og_data'][key] = og.get('content', '')

        # Headings (h1-h3)
        for tag in ['h1', 'h2', 'h3']:
            for h in soup.find_all(tag):
                text = h.get_text(strip=True)
                if text and len(text) > 2:
                    result['headings'].append({'level': tag, 'text': text[:200]})

        # Main text via readability
        try:
            doc = Document(resp.text)
            readable_html = doc.summary()
            readable_soup = BeautifulSoup(readable_html, 'html.parser')
            result['main_text'] = readable_soup.get_text(separator='\n', strip=True)[:5000]
        except:
            paragraphs = soup.find_all('p')
            result['main_text'] = '\n'.join(
                [p.get_text(strip=True) for p in paragraphs if len(p.get_text(strip=True)) > 20]
            )[:5000]

        # Links (top 50)
        seen_hrefs = set()
        for a in soup.find_all('a', href=True):
            if len(result['links']) >= 50:
                break
            text = a.get_text(strip=True)
            href = a['href']
            if href.startswith('/'):
                href = urljoin(url, href)
            if text and href and href not in seen_hrefs and href.startswith('http'):
                seen_hrefs.add(href)
                result['links'].append({'text': text[:100], 'href': href})

        # Images (top 20)
        seen_src = set()
        for img in soup.find_all('img', src=True):
            if len(result['images']) >= 20:
                break
            src = img['src']
            if src.startswith('/'):
                src = urljoin(url, src)
            if src not in seen_src and src.startswith('http'):
                seen_src.add(src)
                result['images'].append({
                    'src': src,
                    'alt': img.get('alt', '')[:100]
                })

        # Tables (top 5)
        for table in soup.find_all('table')[:5]:
            rows = []
            for tr in table.find_all('tr')[:50]:
                cells = [td.get_text(strip=True)[:200] for td in tr.find_all(['td', 'th'])]
                if any(cells):
                    rows.append(cells)
            if rows:
                result['tables'].append(rows)

        # Product-specific extraction (schema.org)
        scripts = soup.find_all('script', type='application/ld+json')
        for script in scripts:
            try:
                ld = json.loads(script.string)
                if isinstance(ld, list):
                    ld = ld[0]
                if isinstance(ld, dict) and ld.get('@type') in ['Product', 'Offer']:
                    result['product_data'] = {
                        'name': ld.get('name', ''),
                        'price': ld.get('offers', {}).get('price', '') if isinstance(ld.get('offers'), dict) else '',
                        'currency': ld.get('offers', {}).get('priceCurrency', '') if isinstance(ld.get('offers'), dict) else '',
                        'description': ld.get('description', '')[:500],
                        'brand': ld.get('brand', {}).get('name', '') if isinstance(ld.get('brand'), dict) else str(ld.get('brand', '')),
                        'rating': str(ld.get('aggregateRating', {}).get('ratingValue', '')) if isinstance(ld.get('aggregateRating'), dict) else '',
                    }
            except:
                pass

    except requests.exceptions.Timeout:
        result['status'] = 'error'
        result['error'] = 'Request timed out after 15 seconds'
    except requests.exceptions.HTTPError as e:
        result['status'] = 'error'
        result['error'] = f"HTTP {e.response.status_code}: {e.response.reason}"
    except Exception as e:
        result['status'] = 'error'
        result['error'] = str(e)[:200]

    return result


# ============================================================
# AI Processing
# ============================================================

def call_groq(text, prompt, api_key, model):
    """Call Groq API for AI processing."""
    try:
        resp = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json"
            },
            json={
                "model": model,
                "messages": [
                    {
                        "role": "system",
                        "content": "You are a precise data extraction assistant. Follow the user's instructions exactly. Return clean, structured text suitable for CSV export."
                    },
                    {
                        "role": "user",
                        "content": f"{prompt}\n\n---\nContent:\n{text[:12000]}"
                    }
                ],
                "temperature": 0.2,
                "max_tokens": 4000
            },
            timeout=60
        )
        resp.raise_for_status()
        data = resp.json()
        return {
            'result': data['choices'][0]['message']['content'],
            'model': model,
            'usage': data.get('usage', {})
        }
    except requests.exceptions.HTTPError as e:
        error_body = e.response.text[:200] if e.response else str(e)
        return {'error': f"Groq API error: {error_body}"}
    except Exception as e:
        return {'error': f"Groq error: {str(e)}"}


def call_gemini(text, prompt, api_key, model):
    """Call Google Gemini API for AI processing."""
    try:
        resp = requests.post(
            f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}",
            json={
                "contents": [{
                    "parts": [{"text": f"{prompt}\n\n---\nContent:\n{text[:12000]}"}]
                }],
                "generationConfig": {
                    "temperature": 0.2,
                    "maxOutputTokens": 4000
                }
            },
            timeout=60
        )
        resp.raise_for_status()
        data = resp.json()
        candidates = data.get('candidates', [])
        if candidates:
            parts = candidates[0].get('content', {}).get('parts', [])
            result_text = parts[0].get('text', '') if parts else ''
            return {'result': result_text, 'model': model}
        return {'error': 'No response from Gemini'}
    except requests.exceptions.HTTPError as e:
        error_body = e.response.text[:200] if e.response else str(e)
        return {'error': f"Gemini API error: {error_body}"}
    except Exception as e:
        return {'error': f"Gemini error: {str(e)}"}


# ============================================================
# Flask Routes
# ============================================================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Parse CSV/Excel to extract URLs."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    filename = file.filename.lower()
    urls = []

    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                for cell in row:
                    if cell and re.search(r'https?://|www\.', cell, re.I):
                        url = cell.strip()
                        if not url.startswith('http'):
                            url = 'https://' + url
                        urls.append(url)

        elif filename.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str) and re.search(r'https?://|www\.', cell, re.I):
                            url = cell.strip()
                            if not url.startswith('http'):
                                url = 'https://' + url
                            urls.append(url)
            wb.close()
        else:
            return jsonify({'error': 'Unsupported file format. Use .csv or .xlsx'}), 400

        urls = list(dict.fromkeys(urls))  # deduplicate preserving order
        return jsonify({'urls': urls, 'count': len(urls)})

    except Exception as e:
        return jsonify({'error': f'File parsing error: {str(e)}'}), 500


@app.route('/api/fetch', methods=['POST'])
def fetch_urls():
    """Fetch and extract content from a list of URLs."""
    data = request.json or {}
    urls = data.get('urls', [])

    if not urls:
        return jsonify({'error': 'No URLs provided'}), 400

    urls = urls[:100]  # Safety limit
    results = []

    def process_url(url):
        url_type = classify_url(url)
        if url_type == 'youtube':
            return extract_youtube(url)
        elif url_type == 'instagram':
            return extract_instagram(url)
        else:
            return extract_general(url)

    with ThreadPoolExecutor(max_workers=5) as executor:
        future_map = {executor.submit(process_url, url): url for url in urls}
        for future in as_completed(future_map):
            url = future_map[future]
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                results.append({
                    'url': url,
                    'type': classify_url(url),
                    'status': 'error',
                    'error': str(e)[:200]
                })

    # Restore original order
    url_order = {url: i for i, url in enumerate(urls)}
    results.sort(key=lambda r: url_order.get(r.get('url', ''), 999))

    return jsonify({'results': results})


@app.route('/api/ai-process', methods=['POST'])
def ai_process():
    """Process text content with AI (Groq or Gemini)."""
    data = request.json or {}
    text = data.get('text', '')
    prompt = data.get('prompt', 'Summarize this content concisely.')
    provider = data.get('provider', 'groq')
    api_key = data.get('api_key', '')
    model = data.get('model', '')

    if not api_key:
        return jsonify({'error': 'API key is required for AI processing'}), 400
    if not text:
        return jsonify({'error': 'No text content to process'}), 400
    if not model:
        return jsonify({'error': 'Please select a model'}), 400

    if provider == 'groq':
        return jsonify(call_groq(text, prompt, api_key, model))
    elif provider == 'gemini':
        return jsonify(call_gemini(text, prompt, api_key, model))
    else:
        return jsonify({'error': f'Unknown provider: {provider}'}), 400


@app.route('/api/ai-batch', methods=['POST'])
def ai_batch():
    """Process multiple items with AI."""
    data = request.json or {}
    items = data.get('items', [])
    prompt = data.get('prompt', 'Summarize this content concisely.')
    provider = data.get('provider', 'groq')
    api_key = data.get('api_key', '')
    model = data.get('model', '')

    if not api_key or not model:
        return jsonify({'error': 'API key and model required'}), 400

    results = []
    for item in items[:20]:  # Limit batch size
        text = item.get('text', '')
        url = item.get('url', '')
        if text:
            if provider == 'groq':
                ai_result = call_groq(text, prompt, api_key, model)
            else:
                ai_result = call_gemini(text, prompt, api_key, model)
            ai_result['url'] = url
            results.append(ai_result)
            time.sleep(0.5)  # Rate limit courtesy

    return jsonify({'results': results})


@app.route('/api/validate-key', methods=['POST'])
def validate_key():
    """Validate an API key and return available models."""
    data = request.json or {}
    provider = data.get('provider', '')
    api_key = data.get('api_key', '')

    if not api_key:
        return jsonify({'valid': False, 'error': 'No API key provided'})

    if provider == 'groq':
        try:
            resp = requests.get(
                "https://api.groq.com/openai/v1/models",
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=10
            )
            if resp.status_code == 200:
                all_models = resp.json().get('data', [])
                # Filter to chat-capable models
                chat_models = sorted(
                    [m['id'] for m in all_models if m.get('id') and 'whisper' not in m['id'].lower()],
                    key=str.lower
                )
                return jsonify({'valid': True, 'models': chat_models})
            return jsonify({'valid': False, 'error': f'HTTP {resp.status_code}'})
        except Exception as e:
            return jsonify({'valid': False, 'error': str(e)})

    elif provider == 'gemini':
        try:
            resp = requests.get(
                f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}",
                timeout=10
            )
            if resp.status_code == 200:
                all_models = resp.json().get('models', [])
                gen_models = sorted([
                    m['name'].replace('models/', '')
                    for m in all_models
                    if 'generateContent' in str(m.get('supportedGenerationMethods', []))
                ])
                return jsonify({'valid': True, 'models': gen_models})
            return jsonify({'valid': False, 'error': f'HTTP {resp.status_code}'})
        except Exception as e:
            return jsonify({'valid': False, 'error': str(e)})

    return jsonify({'valid': False, 'error': 'Unknown provider'})


@app.route('/api/export', methods=['POST'])
def export_csv():
    """Export selected data as CSV download."""
    data = request.json or {}
    rows = data.get('rows', [])
    columns = data.get('columns', [])

    if not rows or not columns:
        return jsonify({'error': 'No data to export'}), 400

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)

    for row in rows:
        csv_row = []
        for col in columns:
            val = row.get(col, '')
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            csv_row.append(str(val) if val else '')
        writer.writerow(csv_row)

    output.seek(0)
    mem = io.BytesIO(output.getvalue().encode('utf-8-sig'))  # BOM for Excel compat
    return send_file(
        mem,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'browser_lite_export_{int(time.time())}.csv'
    )


# ============================================================
# Main
# ============================================================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
